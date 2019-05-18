
import os
import re
import sys
import datetime
import traceback
import itertools
import json

import openpyxl

from datetime			import datetime, date, timedelta
from openpyxl 			import *
from openpyxl			import Workbook
from openpyxl			import load_workbook
from openpyxl.styles	import Color, PatternFill, Font, Border, Side, Alignment, Protection, colors
from openpyxl.comments  import Comment
from openpyxl.worksheet import Worksheet
from openpyxl.utils 	import get_column_letter

class CompetenceUpload (object):
	
	cty_id = []
	skl_lst = []
	ohc_prsn_inf_lst = []
	ohc_prsn_comp_lst = []
	ohc_prsn_trnng_lst = []
	ohc_prsn_trnng_lst_sub = []
	ohc_add_pnts_lst = []
	tc_prsn_inf_lst = []
	new_res_sgl = []
	new_inf_sgl = []
	lst_res_sgl = []
	new_comp_lst = []
	sub_fll_lst = []
	
	def __init__(self):
		super(CompetenceUpload, self).__init__()
		self.get_skl_lst()
		self.get_cty_lst()
		
	def elog(self, nme, msg): 

		'''Creates log file'''

		with open('{path}logs\\log_{nme}.txt'.format(path=path, nme=nme), 'a') as log_file:
			log_file.write('\n{sep}'.format(sep='---------------------------------------------------------------'))
			log_file.write('\n{msg}'.format(msg=msg))
			log_file.write('\n{sep}'.format(sep='---------------------------------------------------------------'))

	def record_sub_data(self, lst, nme):

		try:
			with open('{path}Results\\Sub_Information\\{nme}.txt'.format(path=path, nme=nme), 'a') as file:
				file.write('-------------------------------------------------------------------------------------')
				file.write('\n'+str(datetime.now())+'\n')
				for aa in lst:
					#js = json.dumps(aa, ensure_ascii=False)
					#file.write(js)
					file.write(str(aa))
					file.write(',')
					file.write('\n\n')

		except Exception as e:
			self.elog(
						'record_sub_data',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])
						)
			)

	def empt_skl_fllng(self, id, skl, type):

		try:
			if type == 'max':
				empty_skl = {}
				empty_skl['skl_uid'] = id
				empty_skl['skl_nme'] = skl
				empty_skl['val_su'] = '80'
				empty_skl['val_pmv'] = '80'
				empty_skl['val_um'] = '80'
				empty_skl['training'] = '20'

			if type == 'min':
				empty_skl = {}
				empty_skl['skl_uid']= id
				empty_skl['skl_nme']= skl
				empty_skl['val_su']= '0'
				empty_skl['val_pmv']= '0'
				empty_skl['val_um']= '0'
				empty_skl['training']= '0'

			return empty_skl

		except Exception as e:
			self.elog(
						'empt_skl_fllng',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])
						)
			)
			return {}

	def lod_ohc_prsn_inf(self):
		
		'''Loading prsn Base information from ohc File '''
		try:

			ohc_prsn_inf_lst = []
			max_row = 0
			
			prsn_id_str= 'Persone id'.lower()
			prsn_ru_nme_str= 'ФИО'.lower()
			prsn_en_nme_str= 'Name, Surname'.lower()
			prsn_country_str= 'Country'.lower()
			prsn_cty_str= 'City'.lower()
			prsn_bu_str= 'Bussines Unit'.lower()
			prsn_departament_str= 'Department'.lower()
			prsn_function_str= 'Function'.lower()

			wb_inf = load_workbook(self.ohc_fle_path)
			ws = wb_inf['Persone Base']
			
			#Selecting current column index
			for col_index in range(1,50):
				cur_cell = str(ws.cell(row = 1, column = col_index).value).lower()
				if prsn_id_str == cur_cell: prsn_id_col = col_index
				if prsn_ru_nme_str == cur_cell: prsn_ru_nme_col = col_index
				if prsn_en_nme_str == cur_cell: prsn_en_nme_col = col_index
				if prsn_country_str == cur_cell: prsn_country_col = col_index
				if prsn_cty_str == cur_cell: prsn_cty_col = col_index
				if prsn_bu_str == cur_cell: prsn_bu_col	 = col_index
				if prsn_departament_str == cur_cell: prsn_departament_col= col_index
				if prsn_function_str == cur_cell: prsn_function_col = col_index
				if prsn_target_ur_str == cur_cell: prsn_target_ur_col= col_index
				if prsn_ytd_ur_str	 == cur_cell: prsn_ytd_ur_col = col_index
				
			#Find max row number
			for row_index in range(2, ws.max_row+1):	
				if ws.cell(row=row_index, column=1).value == 'Конец списка':
					max_row = row_index
			
			#Load prsn information
			for row_index in range(2, max_row+1):
				
				if (((str(ws.cell(row=row_index, column=prsn_bu_col).value).lower() == 'PHY BU'.lower()) and
					(ws.cell(row=row_index, column=prsn_function_col).value != 'Manager') and
					(ws.cell(row=row_index, column=prsn_function_col).value != 'PHY Power TCE'))):
					
					select_line = {}
					
					select_line['prsn_id'] = str(ws.cell(row=row_index, column=prsn_id_col).value)
					select_line['prsn_ru_nme'] = str(ws.cell(row=row_index, column=prsn_ru_nme_col).value)
					select_line['prsn_en_nme']= str(ws.cell(row=row_index, column=prsn_en_nme_col).value)
					select_line['prsn_country']= str(ws.cell(row=row_index, column=prsn_country_col).value)
					select_line['prsn_cty'] = str(ws.cell(row=row_index, column=prsn_cty_col).value).replace ("'","")
					select_line['prsn_cty_id']= '0'
					
					for aa in self.cty_id:
						if select_line['prsn_cty'].lower() == aa['eng_cty'].lower() :
							select_line['prsn_cty_id']= aa ['db_id']
							
					select_line['prsn_bu']= str(ws.cell(row=row_index, column=prsn_bu_col).value)
					select_line['prsn_department']= str(ws.cell(row=row_index, column=prsn_departament_col).value)
					select_line['prsn_function']= str(ws.cell(row=row_index, column=prsn_function_col).value)
					select_line['prsn_target_ur']= str(ws.cell(row=row_index, column=prsn_target_ur_col).value)
					select_line['prsn_ur']= str(ws.cell(row=row_index, column=prsn_ytd_ur_col).value)

					if str(ws.cell(row=row_index, column=prsn_id_col).value).lower() == 'id123456'.lower():
						select_line['prsn_function'] = 'UPS prsn'

					for key in select_line:
						if select_line[key] == 'No Data':
							select_line[key] = '0'
					
					ohc_prsn_inf_lst.append(select_line)
			
			for line in ohc_prsn_inf_lst:
				for key in line:
					line[key] = line[key].strip(' ')
					line[key] = line[key].replace ("'", "")
			
			ohc_prsn_inf_lst.sort(key=lambda item: item['prsn_id'])
			self.ohc_prsn_inf_lst = ohc_prsn_inf_lst

			self.record_sub_data(self.ohc_prsn_inf_lst, 'ohc_prsn_inf_lst')

		except Exception as e:
			self.elog(
						'lod_ohc_prsn_inf',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
															d=datetime.now(), type=str(type(e)),
															trb1=traceback.format_tb(sys.exc_info()[2])[0],
															trb2=str(sys.exc_info()[1])
						)
			)
			
	def lod_ohc_prsn_comp(self):
		
		'''Loading Owner Rating information from ohc File '''
		try:
			
			ohc_prsn_comp_lst= []
			
			prsn_id_str= 'prsn id'.lower()
			skl_nme_str= 'Skill Name'.lower()
			val_su_str= 'SU'.lower()
			val_pmv_str= 'PMV'.lower()
			val_um_str= 'UM'.lower()
			
			wb_comp = load_workbook(self.ohc_fle_path)
			ws = wb_comp['Owner Rating']
			
			#Selecting current column index
			for col_index in range(1,50):
				cur_cell = str(ws.cell(row = 1, column = col_index).value).lower()
				if prsn_id_str == cur_cell: prsn_id_col = col_index
				if skl_nme_str == cur_cell: skl_nme_col = col_index
				if val_su_str == cur_cell: val_su_col = col_index
				if val_pmv_str == cur_cell: val_pmv_col = col_index
				if val_um_str == cur_cell: val_um_col = col_index

			for row_index in range (2, ws.max_row+1):	
			
				select_line = {}
				
				select_line['prsn_id'] = str(ws.cell(row = row_index, column = prsn_id_col).value)
				select_line['skl_nme']= str(ws.cell(row = row_index, column = skl_nme_col).value)
				select_line['val_su'] = str(ws.cell(row = row_index, column = val_su_col).value)
				select_line['val_pmv']= str(ws.cell(row = row_index, column = val_pmv_col).value)
				select_line['val_um']= str(ws.cell(row = row_index, column = val_um_col).value)
				
				ohc_prsn_comp_lst.append(select_line)
			
			for line in ohc_prsn_comp_lst:
				for key in line:
					line[key] = line[key].strip(' ')
			
			self.ohc_prsn_comp_lst = ohc_prsn_comp_lst
			self.record_sub_data(self.ohc_prsn_comp_lst, 'ohc_prsn_comp_lst')

		except Exception as e:
			self.elog(
						'lod_ohc_prsn_comp',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2 = str(sys.exc_info()[1])))

	def lod_ohc_prsn_trnng(self):
		
		try:

			ohc_prsn_trnng_lst = []
			max = 0
			n = 1
			comp_str = 'OO_'
			
			first_compet_str= '01'.lower()
			
			wb_trn = load_workbook(self.ohc_fle_path) 
			ws = wb_trn['Training']
			
			for row_index in range (3, ws.max_row):
				if  ws.cell(row = row_index+1 , column = 1).value is None :
					max = int(row_index) - 1
			
			#Selecting current column index
			for col_index in range(1,50):
				cur_cell = str(ws.cell(row = 1, column = col_index).value).lower()
				if first_compet_str 		in cur_cell : first_compet_col = col_index
			
			for row_index in range (4, max+1):
						
				select_line = {}
				
				select_line['num'] = '{n}'.format (n = n)
				select_line['prsn_id'] = str(ws.cell(row = row_index, column = 1).value)
			
				for column_index in range(first_compet_col, 200):
				
					val = str(ws.cell(row = 1, column = column_index).value)
					
					if comp_str in val:
					
						competence = str(ws.cell(row = 1, column = column_index).value)
						select_line['{comp}'.format(comp = competence)] = str(ws.cell(
																						row=row_index,
																						column=column_index).value)
		
				ohc_prsn_trnng_lst.append(select_line)
				n = n+1
			
			for line in ohc_prsn_trnng_lst:
				for key in line:
					line[key] = line[key].strip(' ')
			
			ohc_prsn_trnng_lst_sub  = []
			
			
			for line in ohc_prsn_trnng_lst:
			
				sub_lst = line
				for key in line:
					if (line[key] == u'#Н/Д') or (line[key] == 'No Data') or (line[key] == '#N/A') or \
						(line[key] == 'None') or (line[key] is None) :
						sub_lst[key] = '0'
			
				ohc_prsn_trnng_lst_sub.append(sub_lst)
			
			self.ohc_prsn_trnng_lst = ohc_prsn_trnng_lst_sub
			self.record_sub_data(self.ohc_prsn_trnng_lst, 'ohc_prsn_trnng_lst')

		except Exception as e:
			self.elog(
						'lod_ohc_prsn_trnng',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
															d=datetime.now(), type=str(type(e)),
															trb1=traceback.format_tb(sys.exc_info()[2])[0],
															trb2=str(sys.exc_info()[1])
						)
			)

	def lod_ohc_add_pnts(self):

		try:
			ohc_add_pnts_lst= []

			skl_nme_str= 'Skill Name'.lower()
			prsn_id_str= 'prsn id'.lower()
			ad_val_su_str= 'Количество часов SU'.lower()
			ad_val_pmv_str= 'Количество часов PMV'.lower()
			ad_val_um_str= 'Количество часов UM'.lower()

			wb_add = load_workbook(self.ohc_fle_path)
			ws = wb_add['Template All']

			#Selecting current column index
			for col_index in range(1,50):
				cur_cell=str(ws.cell(row=1, column=col_index).value).lower()
				if prsn_id_str == cur_cell: prsn_id_col = col_index
				if skl_nme_str == cur_cell: skl_nme_col = col_index
				if ad_val_su_str == cur_cell: ad_val_su_col = col_index
				if ad_val_pmv_str == cur_cell: ad_val_pmv_col = col_index
				if ad_val_um_str == cur_cell: ad_val_um_col = col_index

			for row_index in range (2, ws.max_row+1):

				select_line = {}

				select_line['prsn_id']=str(ws.cell(row=row_index, column=prsn_id_col).value)
				select_line['skl_nme']= str(ws.cell(row=row_index, column=skl_nme_col).value)
				select_line['ad_val_su']=str(ws.cell(row=row_index, column=ad_val_su_col).value)
				select_line['ad_val_pmv']= str(ws.cell(row=row_index, column=ad_val_pmv_col).value)
				select_line['ad_val_um']= str(ws.cell(row=row_index, column=ad_val_um_col).value)

				ohc_add_pnts_lst.append(select_line)

			for line in ohc_add_pnts_lst:
				for key in line:
					line[key] = line[key].strip(' ')
			
			self.ohc_add_pnts_lst = ohc_add_pnts_lst
			self.record_sub_data(self.ohc_add_pnts_lst, 'ohc_add_pnts_lst')

		except Exception as e:
			self.elog(
						'lod_ohc_add_pnts',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
															d=datetime.now(), type=str(type(e)),
															trb1=traceback.format_tb(sys.exc_info()[2])[0],
															trb2=str(sys.exc_info()[1])
						)
			)

	def lod_tc(self):
		
		try:
			
			tc_prsn_inf_lst = []
			comp_str = 'OO_'.lower()
			
			tc_theme_str= 'Theme'.lower()
			tc_date_str= 'M'.lower()
			tc_customer_str= 'Customer'.lower()
			prsn_nme_str= 'Name'.lower()
			duration_str= 'Durat.'.lower()

			wb_inf = load_workbook(self.tc_fle_path)
			ws = wb_inf['Sheet1']
			
			#Selecting current column index
			for col_index in range(1,50):
				cur_cell = str(ws.cell(row = 1, column = col_index).value).lower()
				if tc_theme_str == cur_cell: tc_theme_col = col_index
				if tc_date_str == cur_cell: tc_date_col = col_index
				if tc_customer_str == cur_cell: tc_customer_col = col_index
				if prsn_nme_str == cur_cell: prsn_nme_col = col_index
				if duration_str == cur_cell: duration_col = col_index

			#Load TC information
			for row_index in range(2, ws.max_row+1):
				if month == '12':
					next_month = {}
					next_month['year'] = int(year)+1
					next_month['month'] = 1
				else:
					next_month = {}
					next_month['year'] = int(year)
					next_month['month'] = int(month)+1

				if ((comp_str in str(ws.cell(row = row_index, column = tc_theme_col).value).lower())and
					(date(2000+int(str(ws.cell(row=row_index, column=tc_date_col).value)[6:8]),
						   int(str(ws.cell(row=row_index, column=tc_date_col).value)[3:5]),
						   int(str(ws.cell(row=row_index, column=tc_date_col).value)[0:2])) +
					 timedelta(days = int(ws.cell(row=row_index, column=duration_col).value))
					 	< date(next_month['year'],next_month['month'] ,1)) and
					(ws.cell(row = row_index, column = tc_customer_col).value == 'OHC Company')):
					
					select_line = {}
					
					select_line['tc_theme'] = str(ws.cell(row=row_index, column=tc_theme_col).value).strip(' ')
					select_line['tc_month']= int(str(ws.cell(row=row_index, column=tc_date_col).value)[4:5])
					select_line['tc_customer']= str(ws.cell(row=row_index, column=tc_customer_col).value).strip(' ')
					select_line['prsn_nme']= str(ws.cell(row=row_index, column=prsn_nme_col).value).strip(' ')

					tc_prsn_inf_lst.append(select_line)
			
			self.tc_prsn_inf_lst = tc_prsn_inf_lst
			self.record_sub_data(self.tc_prsn_inf_lst, 'tc_prsn_inf_lst')

		except Exception as e:
			self.elog(
						'lod_tc',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])))
	
	def lod_lst_fle(self):

		try:
			lst_res_sgl= []
			wb_lst = load_workbook(self.lst_fle_path)
			ws = wb_lst['prsn_skl']

			#Find max row number
			for row_index in range(2, ws.max_row+1):
				if ws.cell(row = row_index, column = 1).value is None:
					max_row = row_index - 1

			for column_index in range(1, 100):
				if ws.cell(row = 2, column = column_index).value == 'SKL_VOL_PER_prsn':
					max_column = column_index - 2

			for row_index in range(3, max_row - 1):

				if ws.cell(row=row_index,column=7).value  is not None:

					select_line = {}

					select_line['prsn_uid'] = str(ws.cell(row=row_index,column=1).value)
					select_line['prsn_nme_en'] = str(ws.cell(row=row_index,column=5).value)
					select_line['skl_lst'] = []

					for column_index in range(8,max_column+1):
						if ws.cell(row=row_index, column=column_index).value is not None:
							select_line['skl_lst'].append(str(ws.cell(row=row_index, column=column_index).value))

					lst_res_sgl.append(select_line)

			self.lst_res_sgl = lst_res_sgl
			self.record_sub_data(self.lst_res_sgl, 'lst_res_sgl')

		except Exception as e:
			self.elog(
						'lod_lst_fle',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
															d=datetime.now(), type=str(type(e)),
															trb1=traceback.format_tb(sys.exc_info()[2])[0],
															trb2=str(sys.exc_info()[1])
						)
			)

	def add_trnng (self, prsn_id, prsn_nme, skl):
		
		try:

		    #Part 1 - From ohc
			indicator = 0
			training = ''
			
			#print (ohc_prsn_trnng_lst)
			
			for line_t in self.ohc_prsn_trnng_lst_sub:
				for key in line_t:
				
					if (prsn_id.lower() == line_t['prsn_id'].lower()) and (skl.lower() == key.lower()):
							
						if line_t[key] != '0':
					
							indicator = indicator+1
							training = '16'														
				
					if indicator == 0 :
						training = '0'

			#Part 2 - From TC
			for line_t in self.tc_prsn_inf_lst:
				if (prsn_nme.lower() == line_t['prsn_nme'].split(' ')[0].lower()) and (skl.lower() ==
																						line_t['tc_theme'].lower()):

						indicator = indicator+1
						training = '16'

				if indicator == 0 :
					training = '0'

			return training

		except Exception as e:
			self.elog('add_trnng','{d}_{type}\n{trb1}\n{trb2} '.format(d=datetime.now(), type=str(type(e)),
						trb1=traceback.format_tb(sys.exc_info()[2])[0], trb2=str(sys.exc_info()[1])))
			return ''

	def cmpr_sub_fll_lst(self):

		try:

			sub_fll_lst = []

			for line_i in self.ohc_prsn_inf_lst:

				one_line = {}
				skl_dict= []

				prsn_id  = str(line_i['prsn_id'])
				prsn_nme= line_i['prsn_ru_nme'].split(' ')[0]

				one_line['prsn_uid'] = line_i['prsn_id']
				one_line['prsn_nme_ru'] = line_i['prsn_ru_nme']
				one_line['prsn_nme_en'] = line_i['prsn_en_nme']
				one_line['prsn_country'] = line_i['prsn_country']
				one_line['prsn_src_nme'] = line_i['prsn_cty']
				one_line['prsn_src_uid'] = line_i['prsn_cty_id']
				one_line['prsn_function'] = line_i['prsn_function']

				for line_g in self.ohc_prsn_comp_lst:
					for skls in self.skl_lst:

						if (((line_i['prsn_id']).lower() == (line_g['prsn_id']).lower()) and
							((skls['skl']).lower() == (line_g['skl_nme']).lower())):

								skl = {}

								skl_n = str(line_g['skl_nme'])

								skl['skl_uid'] = skls['id']
								skl['skl_nme'] = line_g['skl_nme']
								skl['val_su'] = line_g['val_su']
								skl['val_pmv'] = line_g['val_pmv']
								skl['val_um'] = line_g['val_um']

								training = self.add_trnng(prsn_id, prsn_nme, skl_n)

								skl['training']= training

								skl_dict.append(skl)

				res = [skl['skl_nme'] for skl in skl_dict]

				try:
					for skls in self.skl_lst:
						if skls['skl'] not in res:

							id = skls['id']
							skl = skls['skl']
							empty_skl = {}

							if (skls['skl'].lower() == 'OO_PHY_C_ABSTRACT'.lower()):
								#print (line_i['prsn_function'].lower())
								if ('cooling' in (line_i['prsn_function']).lower()):

									empty_skl = self.empt_skl_fllng(id, skl, 'max')

								if (('ups' in (line_i['prsn_function']).lower()) or
										((line_i['prsn_function']).lower() == 'networking prsn')):

									empty_skl = self.empt_skl_fllng(id, skl, 'min')


							elif (skls['skl'].lower() == 'OO_PHY_P_ABSTRACT'.lower()):

								if (((line_i['prsn_function']).lower() == 'cooling prsn') or
									((line_i['prsn_function']).lower() == 'cooling team leader') or
									((line_i['prsn_function']).lower() == 'networking prsn')):

									empty_skl = self.empt_skl_fllng(id, skl, 'min')

								if (((line_i['prsn_function']).lower() == 'ups prsn') or
									((line_i['prsn_function']).lower() == 'ups team leader')):

									empty_skl = self.empt_skl_fllng(id, skl, 'max')


							elif (skls['skl'].lower() == 'OO_PHY_ANY_ABSTRACT'.lower()):

								if (((line_i['prsn_function']).lower() == 'cooling prsn') or
									((line_i['prsn_function']).lower() == 'cooling team leader') or
									((line_i['prsn_function']).lower() == 'ups prsn') or
									((line_i['prsn_function']).lower() == 'ups team leader')):

									empty_skl = self.empt_skl_fllng(id, skl, 'max')

								if (line_i['prsn_function']).lower() == 'networking prsn':

									empty_skl = self.empt_skl_fllng(id, skl, 'min')


							elif (skls['skl'].lower() == 'OO_PHY_B_ABSTRACT'.lower()):

								if (((line_i['prsn_function']).lower() == 'cooling prsn') or
									((line_i['prsn_function']).lower() == 'cooling team leader')):

									empty_skl = self.empt_skl_fllng(id, skl, 'min')

								if (((line_i['prsn_function']).lower() == 'ups prsn') or
									((line_i['prsn_function']).lower() == 'ups team leader') or
									((line_i['prsn_function']).lower() == 'networking prsn')):

									empty_skl = self.empt_skl_fllng(id, skl, 'max')

							else:

								empty_skl['skl_uid']= skls['id']
								empty_skl['skl_nme']= skls['skl']
								empty_skl['val_su']= '0'
								empty_skl['val_pmv']= '0'
								empty_skl['val_um']= '0'

								skl_n = str(skls['skl'])

								training = self.add_trnng(prsn_id, prsn_nme, skl_n)

								empty_skl['training']= training

						skl_dict.append(empty_skl)

						skl_dict.sort(key = lambda item : int(item['skl_uid']))

				except Exception as e :
					self.elog(
								'empty_skls',
								'{d}_{type}\n{trb1}\n{trb2} '.format(
																	d=datetime.now(), type=str(type(e)),
																	trb1=traceback.format_tb(sys.exc_info()[2])[0],
																	trb2=str(sys.exc_info()[1])
								)
					)

				for line_s in skl_dict:
					for line_a in self.ohc_add_pnts_lst:

						if ((line_s['skl_nme'].lower() == line_a['skl_nme'].lower()) and
							(one_line['prsn_uid']).lower() == line_a['prsn_id'].lower()):
							line_s['val_su'] = str(int(line_s['val_su'])+int(line_a['ad_val_su']))
							line_s['val_pmv'] = str(int(line_s['val_pmv'])+int(line_a['ad_val_pmv']))
							line_s['val_um'] = str(int(line_s['val_um'])+int(line_a['ad_val_um']))

				one_line['prsn_skl_lst'] = skl_dict

				sub_fll_lst.append(one_line)

			self.sub_fll_lst = sub_fll_lst
			self.record_sub_data(self.sub_fll_lst, 'sub_fll_lst')

		except Exception as e:
			self.elog(
						'cmpr_sub_fll_lst',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
															d=datetime.now(), type = str(type(e)),
															trb1 = traceback.format_tb(sys.exc_info()[2])[0],
															trb2 = str(sys.exc_info()[1])
						)
			)

	def mke_prsn_skl(self):

		try:
			prsn_sgl_lst = []

			for prsn in self.sub_fll_lst:

				prsn_uid = prsn['prsn_uid']
				skl_lst = prsn['prsn_skl_lst']
				new_skl_lst = []

				for skl_id in range(1, len(self.skl_lst)+1):
					skl_line= [ skl for skl in skl_lst if skl['skl_uid'] == str('{}'.format(skl_id)) ]

					if (((int(skl_line[0]['val_pmv'])+int(skl_line[0]['training'])) > 40) or
						(((int(skl_line[0]['val_su'])+int(skl_line[0]['val_pmv'])
						+ int(skl_line[0]['val_um'])+int(skl_line[0]['training'])* 3)) / 3) > 50):

						new_skl_lst.append(str('{}'.format(skl_id)))

				prsn_sgl_lst.append({ 'prsn_uid' : prsn_uid, 'skl_lst' : new_skl_lst,  })

			self.new_res_sgl = prsn_sgl_lst
			self.record_sub_data(self.new_res_sgl, 'new_res_sgl')

		except Exception as e:
			self.elog(
						'mke_prsn_skl',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
                                                               d=datetime.now(), type=str(type(e)),
									                           trb1=traceback.format_tb(sys.exc_info()[2])[0],
                                                               trb2=str(sys.exc_info()[1])))
		return []

	def mke_prsn_inf(self):

		prsn_inf = []
		abs_str = 'Abstract prsn'

		try:
			for prsn in self.sub_fll_lst:

				for line_l in self.lst_res_sgl:
					if prsn['prsn_uid'].lower() == line_l['prsn_uid'].lower():

						prsn_inf.append({
							'prsn_uid'		: prsn['prsn_uid'] ,
							'prsn_nme_ru'	: prsn['prsn_nme_ru'] ,
							'prsn_nme_en'	: prsn['prsn_nme_en'] ,
							'prsn_src_nme'	: prsn['prsn_src_nme'] ,
							'prsn_src_uid'	: prsn['prsn_src_uid'] ,
							'prsn_function'	: prsn['prsn_function'] ,
							'prsn_clc_id'	: line_l['prsn_nme_en'],
						})

			prsn_inf.append(
				{'prsn_uid': 'id000001', 'prsn_nme_ru': abs_str, 'prsn_nme_en': abs_str,'prsn_src_nme': 'Mohcow',
				'prsn_src_uid': '1', 'prsn_function': 'Abstract Function', 'prsn_clc_id': 'MSK_prsn',})
			prsn_inf.append(
				{'prsn_uid': 'id000002', 'prsn_nme_ru': abs_str, 'prsn_nme_en': abs_str,'prsn_src_nme': 'Mohcow',
				'prsn_src_uid': '1', 'prsn_function': 'Abstract Function', 'prsn_clc_id': 'MSK_prsn',})
			prsn_inf.append(
				{'prsn_uid': 'id000003', 'prsn_nme_ru': abs_str, 'prsn_nme_en': abs_str,'prsn_src_nme': 'Mohcow',
				'prsn_src_uid': '1', 'prsn_function': 'Abstract Function', 'prsn_clc_id': 'MSK_prsn',})
			prsn_inf.append(
				{'prsn_uid': 'id000004', 'prsn_nme_ru': abs_str, 'prsn_nme_en': abs_str,'prsn_src_nme': 'Mohcow',
				'prsn_src_uid': '1', 'prsn_function': 'Abstract Function', 'prsn_clc_id': 'MSK_prsn',})

			prsn_inf.sort(key=lambda item: item['prsn_nme_ru'])

			self.new_prsn_inf = prsn_inf
			self.record_sub_data(self.new_prsn_inf, 'new_prsn_inf')

		except Exception as e:
			self.elog(
						'mke_prsn_inf',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1]))
															)
			return []

	def mke_sptl_inf_txt(self):

		prsn_sgl = []
		prsn_ohc = []

		try:
			#prsn_sgl
			for prsn in self.new_comp_lst:
				i = '1'
				prsn_sgl_lne = {}
				prsn_sgl_lne['prsn_uid'] = prsn['prsn_uid']
				prsn_sgl_lne['skl_lst'] = []

				for skl_num in range (1, 59):
					if str(skl_num) in prsn['skl_lst']:
						prsn_sgl_lne['skl_lst'].append(str(skl_num))
					else:
						if skl_num < 10:
							prsn_sgl_lne['skl_lst'].append('_')
						if skl_num >= 10:
							prsn_sgl_lne['skl_lst'].append('__')

				prsn_sgl.append(prsn_sgl_lne)

			#prsn_ohc
			for prsn_num in self.new_comp_lst:
				for prsn in self.new_prsn_inf:
					if prsn_num['prsn_uid'] == prsn['prsn_uid']:
						prsn_ohc_line = {}
						prsn_ohc_line['prsn_uid'] = prsn['prsn_uid']
						prsn_ohc_line['function'] = prsn['prsn_function']
						prsn_ohc_line['prsn_src_uid'] = prsn['prsn_src_uid']
						prsn_ohc_line['prsn_src_nme'] = prsn['prsn_src_nme']
						prsn_ohc_line['_prsn_nme_en'] = prsn['prsn_nme_en']
						prsn_ohc_line['prsn_nme_ru'] = prsn['prsn_nme_ru']
						prsn_ohc_line['prsn_nme_en'] = prsn['prsn_clc_id']
						prsn_ohc.append(prsn_ohc_line)

			now = datetime.now()
			date = '{y}.{m}.{d}'.format(y=now.year, m=now.month, d=now.day)
			with open('{path}Results\\txt\\prsn_skl_{date}.txt'.format(path=path, date = date) , 'w') as file:
				file.write('prsn_ohc = ['+'\n')
				for aa in prsn_ohc:
					file.write(str(aa))
					file.write(',\n')
				file.write(']'+'\n')

				file.write('prsn_sgl = ['+'\n')
				sub_index = 0
				for aa in prsn_sgl:
					file.write(str(aa))
					file.write(',#{skl_lst_len}\t-\t{prsn_lctn}\t - {prsn_en_nme}\n'.format(
						skl_lst_len=self.new_comp_lst[sub_index]['skl_lst_len'],
						prsn_lctn=prsn_ohc[sub_index]['prsn_src_nme'],
						prsn_en_nme=prsn_ohc[sub_index]['_prsn_nme_en']
					)
					)
					sub_index += 1
				file.write(']'+'\n')

		except Exception as e:
			self.elog(
						'mke_sptl_inf_txt',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])
						)
			)
			return []

	def cnctnt_lsts(self):

		try:

			new_comp_lst = []
			with open('{path}logs\\renewal.txt'.format(path=path), 'a') as file:
				file.write('-------------------------------------------------------------------------------------')
				file.write('\n'+str(datetime.now())+'\n')

				for line_r in self.lst_res_sgl:
					indicator = 0
					if ((line_r['prsn_uid'].lower() == 'id000001'.lower()) or
							line_r['prsn_uid'].lower() == 'id000002'.lower() or
							(line_r['prsn_uid'].lower() == 'id000003'.lower()) or
									line_r['prsn_uid'].lower() == 'id000004'.lower()):
						indicator += 1

					line_new = {}
					line_new['prsn_uid'] = line_r['prsn_uid']
					line_new['skl_lst'] = line_r['skl_lst']

					for line_l in self.new_res_sgl:
						if line_r['prsn_uid'].lower() == line_l['prsn_uid'].lower():
							indicator += 1
							for element_l in line_l['skl_lst']:
								if element_l not in line_new['skl_lst']:
									msg = 'id: '+str(line_new['prsn_uid'])+', competence: '\
										  +str(element_l)+' - '+ str(datetime.now())
									file.write(msg+'\n')

									line_new['skl_lst'].append(element_l)

					line_new['skl_lst'].sort(key=lambda item: int(item))
					line_new['skl_lst_len'] = len(line_new['skl_lst'])
					if indicator > 0:
						new_comp_lst.append(line_new)
					else:
						print (line_r['prsn_uid']+ u' not in ohc')

			#for line in new_comp_lst:
				#print (line)

			new_comp_lst.sort(key=lambda item: item['skl_lst_len'], reverse=True)
			self.new_comp_lst = new_comp_lst
			self.record_sub_data(self.new_comp_lst, 'new_comp_lst')

		except Exception as e:
			self.elog(
						'cnctnt_lsts',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])
						)
			)

	def mke_fnl_xlsx(self):

		try:

			skl_index = 1
			wb = Workbook()
			ws = wb.active
			ws.title = 'prsn_skl'

			# Set width on columns
			ws.column_dimensions['A'].width = 5.75
			ws.column_dimensions['B'].width = 5.75
			ws.column_dimensions['C'].width = 5.75
			ws.column_dimensions['D'].width = 5.75
			ws.column_dimensions['E'].width = 5.75
			ws.column_dimensions['F'].width = 5.75
			ws.column_dimensions['G'].width = 35

			for column_index in range(8, 11+len(self.skl_lst)):
				ws.column_dimensions['{}'.format(get_column_letter(column_index))].width = 5

			ws.row_dimensions[1].height = 15
			ws.row_dimensions[2].height = 150

			for row_index in range(3, 6+len(self.new_comp_lst)):
				ws.row_dimensions[row_index].height = 15

			ws['A2'].value = 'prsn_uid'
			ws['B2'].value = 'prsn_src_uid'
			ws['C2'].value = 'prsn_src_nme'
			ws['D2'].value = '_prsn_nme_en'
			ws['E2'].value = 'prsn_nme_en'
			ws['F2'].value = 'prsn_function'
			ws['G2'].value = 'prsn_nme_ru'
			ws['{col}2'.format(col = get_column_letter(9+len(self.skl_lst)))] = 'SKL_VOL_PER_prsn'

			for column_index in range(8, 8+len(self.skl_lst)):
				ws.cell(row=1, column=column_index).value = skl_index
				ws.cell(row=1, column=column_index).alignment = Alignment(horizontal='center')

				ws.cell(row=2, column =column_index).value = self.skl_lst[column_index-8]['skl']
				row_f = 4+len(self.new_prsn_inf)
				ws.cell(row=row_f, column=column_index).value = '= COUNT({column}3 : {column}{row_end})'.format(
																								column=get_column_letter(column_index),
																								row_end=row_f - 2)

				ws.cell(row=row_f, column=column_index).data_type = 'f'
				ws.cell(row=row_f, column=column_index).alignment = Alignment(horizontal='center')
				ws.cell(row=row_f, column=column_index).fill = PatternFill(start_color = 'A9D08E', end_color = 'A9D08E', fill_type = 'solid')
				skl_index += 1

			row_index = 3
			for line_i in self.new_prsn_inf:
				ws.cell(row = row_index, column=1).value = line_i['prsn_uid']
				ws.cell(row = row_index, column=2).value = int(line_i['prsn_src_uid'])
				ws.cell(row = row_index, column=3).value = line_i['prsn_src_nme']
				ws.cell(row = row_index, column=4).value = line_i['prsn_nme_en']
				ws.cell(row = row_index, column=5).value = line_i['prsn_clc_id']
				ws.cell(row = row_index, column=6).value = line_i['prsn_function']
				ws.cell(row = row_index, column=7).value = line_i['prsn_nme_ru']
				column_f = 9+len(self.skl_lst)
				ws.cell(row=row_index, column=column_f).value = '= COUNT(H{row} : {end_col}{row})'.format(
																								row=row_index,
																								end_col=get_column_letter(column_f - 2))
				ws.cell(row=row_index, column=column_f).data_type = 'f'
				ws.cell(row=row_index, column=column_f).alignment = Alignment(horizontal='center')
				ws.cell(row=row_index, column=column_f).fill = PatternFill(start_color = 'A9D08E', end_color = 'A9D08E', fill_type = 'solid')

				for line_c in self.new_comp_lst :
					if line_i['prsn_uid'].lower() == line_c['prsn_uid'].lower() :
						for skl in line_c['skl_lst']:
							ws.cell(row=row_index, column=7+int(skl)).value = int(skl)
							ws.cell(row=row_index, column=7+int(skl)).alignment = Alignment(horizontal='center')
				row_index += 1

			ws['G{row}'.format(row = 4+len(self.new_prsn_inf))] = 'prsn_VOL_PER_SKL'

			for index in range(8, 10+len(self.skl_lst)):
				cell = ws['{col}2'.format(col = get_column_letter(index))]
				cell.alignment = Alignment(text_rotation=90,horizontal='center')

			for row_index in range(1, 6+len(self.new_comp_lst)):
				for column_index in range(1,7):
					ws.cell(row= row_index, column =column_index).font = Font(color='A6A6A6')

			ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
			ws['G2'].alignment = Alignment(horizontal='center', vertical='center')

			ws.auto_filter.ref = 'A2:{col}{row}'.format(col = get_column_letter(column_f), row = row_f)
			ws.freeze_panes = ws['H3']

			now = datetime.now()
			date = '{y}.{m}.{d}'.format(y=now.year, m=now.month, d=now.day)
			wb.save('{path}\\Results\\xlsx\\Final_{date}.xlsx'.format(path=path, date = date))

		except Exception as e:
			self.elog(
						'make_final_xlsx',
						'{d}_{type}\n{trb1}\n{trb2} '.format(
																d=datetime.now(), type=str(type(e)),
																trb1=traceback.format_tb(sys.exc_info()[2])[0],
																trb2=str(sys.exc_info()[1])))

	def launch (self, period, lst_file_dte):

		try:

			self.ohc_fle_path = '{path}ohc_{period}.xlsx'.format(path = path, period = period)
			self.tc_fle_path = '{path}tr_{period}.xlsx'.format(path = path, period = period)
			self.lst_fle_path = '{path}prsn_skl_{dte}.xlsx'.format(path = path, dte = lst_file_dte)
			self.period = period

			self.lod_ohc_prsn_inf()
			self.lod_ohc_prsn_comp()
			self.lod_ohc_prsn_trnng()
			self.lod_ohc_add_pnts()
			self.lod_lst_fle()
			self.lod_tc()

			self.cmpr_sub_fll_lst()
			self.mke_prsn_skl()
			self.mke_prsn_inf()
			self.cnctnt_lsts()
			self.mke_sptl_inf_txt()

			self.mke_fnl_xlsx()
		except Exception as e:
			self.elog(
						'launch',
						'{d}_{type}\n{trb1}\n{trb2}'.format(
																d = datetime.now(), type = str(type(e)),
																trb1 = traceback.format_tb(sys.exc_info()[2])[0],
																trb2 = str(sys.exc_info()[1])
						)
			)
			return []
			
path = ''
month = '1'
year = '2019'
period = month+'_'+year
lst_file_dte = '20180101'

# For launght
#x = CompetenceUpload()
#x.launch(period, lst_file_dte)



